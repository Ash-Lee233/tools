"""
PyTorch API Operator Classification Tool

Enumerates all ATen operators from local PyTorch installation and classifies them
into 4 categories based on implementation approach. Results are saved as Excel.

Categories:
  1 - CUDA直接提供算子: Operators using cuBLAS/cuDNN/cuFFT/cuSOLVER libraries
  2 - CUDA C++自定义算子: Operators with custom CUDA C++ kernels
  3 - 算子拼接: Composite operators built from other operators
  4 - Triton开发: Operators with Triton implementations
"""

import torch
import torch._C
import inspect
import re
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


# ============================================================
# 1. CUDA Library Whitelist (cuBLAS, cuDNN, cuFFT, cuSOLVER, cuSPARSE, cuRAND)
# ============================================================

# cuBLAS ops: BLAS routines (matrix multiply, gemm, etc.)
CUBLAS_OPS = {
    "mm", "bmm", "addmm", "addbmm", "baddbmm",
    "matmul", "dot", "vdot", "mv", "ger",
    "addmv", "addr",
    "_trilinear",
    "linalg_multi_dot",
}

# cuDNN ops: convolution, pooling, batchnorm, RNN, etc.
CUDNN_OPS = {
    "cudnn_convolution", "cudnn_convolution_transpose",
    "cudnn_batch_norm", "cudnn_batch_norm_backward",
    "miopen_convolution", "miopen_convolution_transpose",
    "miopen_batch_norm", "miopen_batch_norm_backward",
    "miopen_depthwise_convolution",
    "cudnn_affine_grid_generator", "cudnn_affine_grid_generator_backward",
    "cudnn_grid_sampler", "cudnn_grid_sampler_backward",
    "cudnn_is_acceptable",
    "_cudnn_rnn", "_cudnn_rnn_backward",
    "_cudnn_init_dropout_state",
    "cudnn_convolution_relu", "cudnn_convolution_add_relu",
    # Native conv that dispatches to cuDNN
    "_conv_depthwise2d",
    "slow_conv3d", "slow_conv_dilated2d", "slow_conv_dilated3d",
    "slow_conv_transpose2d", "slow_conv_transpose3d",
    "thnn_conv2d",
}

# cuFFT ops
CUFFT_OPS = {
    "fft_fft", "fft_ifft", "fft_fft2", "fft_ifft2",
    "fft_fftn", "fft_ifftn", "fft_rfft", "fft_irfft",
    "fft_rfft2", "fft_irfft2", "fft_rfftn", "fft_irfftn",
    "fft_hfft", "fft_ihfft", "fft_hfft2", "fft_ihfft2",
    "fft_hfftn", "fft_ihfftn",
    "fft_fftshift", "fft_ifftshift",
    "fft_fftfreq", "fft_rfftfreq",
    "_fft_r2c", "_fft_c2r", "_fft_c2c",
    "stft", "istft",
}

# cuSOLVER ops: linear algebra solvers
CUSOLVER_OPS = {
    "linalg_cholesky", "linalg_cholesky_ex",
    "cholesky", "cholesky_inverse", "cholesky_solve",
    "linalg_inv", "linalg_inv_ex", "inverse",
    "linalg_solve", "linalg_solve_ex",
    "linalg_solve_triangular",
    "linalg_lstsq",
    "linalg_lu", "linalg_lu_factor", "linalg_lu_factor_ex", "linalg_lu_solve",
    "lu_solve", "lu_unpack", "_lu_with_info",
    "linalg_qr", "geqrf",
    "linalg_svd", "linalg_svdvals", "svd",
    "linalg_eig", "linalg_eigh", "linalg_eigvals", "linalg_eigvalsh",
    "_linalg_eigh", "symeig",
    "linalg_det", "linalg_slogdet", "logdet", "det", "slogdet",
    "linalg_matrix_rank", "matrix_rank",
    "linalg_pinv", "pinverse",
    "linalg_norm", "linalg_vector_norm", "linalg_matrix_norm",
    "linalg_cond",
    "linalg_matrix_power", "matrix_power",
    "linalg_matrix_exp", "matrix_exp",
    "linalg_cross",
    "triangular_solve",
    "ormqr",
    "linalg_householder_product",
    "linalg_tensorinv", "linalg_tensorsolve",
    "_linalg_svd",
}

# cuSPARSE ops
CUSPARSE_OPS = {
    "_sparse_mm", "_sparse_sparse_matmul",
    "_sparse_addmm",
    "sparse_sampled_addmm",
    "_sparse_csr_tensor_unsafe",
    "_sparse_coo_tensor_unsafe",
    "_sparse_sum",
    "hspmm",
    "sspaddmm",
    "smm",
}

# Combine all CUDA library ops
CUDA_LIBRARY_OPS = CUBLAS_OPS | CUDNN_OPS | CUFFT_OPS | CUSOLVER_OPS | CUSPARSE_OPS


def normalize_op_name(name: str) -> str:
    """Extract the base operator name from aten::name.overload format."""
    # Remove aten:: prefix
    if name.startswith("aten::"):
        name = name[6:]
    # Remove overload suffix
    if "." in name:
        name = name.split(".")[0]
    return name


def is_cuda_library_op(base_name: str) -> bool:
    """Check if an operator is backed by a CUDA library (cuBLAS/cuDNN/cuFFT/cuSOLVER)."""
    return base_name in CUDA_LIBRARY_OPS


# ============================================================
# 2. Enumerate all ATen operators
# ============================================================

def get_all_aten_ops():
    """Get all aten operators from JIT schemas."""
    schemas = torch._C._jit_get_all_schemas()
    aten_ops = []
    seen = set()

    for schema in schemas:
        name = str(schema.name)
        overload = schema.overload_name
        if overload:
            full_name = f"aten::{name}.{overload}"
        else:
            full_name = f"aten::{name}"

        # Only keep aten namespace
        if not name.startswith("aten::") and "aten" not in str(schema):
            # The schema name might not have the namespace prefix;
            # check the actual namespace
            pass

        # Filter: schema.name gives just the op name without namespace for some versions
        # We need to reconstruct from the full schema string
        schema_str = str(schema)
        if not schema_str.startswith("aten::"):
            continue

        # Parse full name from schema string
        match = re.match(r'(aten::\S+)\(', schema_str)
        if not match:
            continue
        full_name = match.group(1)

        if full_name in seen:
            continue
        seen.add(full_name)

        aten_ops.append({
            "full_name": full_name,
            "schema": schema,
            "schema_str": schema_str,
        })

    return aten_ops


# ============================================================
# 3. Dispatch key detection
# ============================================================

def get_op_dispatch_info(op_name: str):
    """
    Get dispatch information for an operator.
    Returns dict with keys: is_composite_implicit, is_composite_explicit,
    has_cuda_kernel, has_cpu_kernel
    """
    info = {
        "is_composite_implicit": False,
        "is_composite_explicit": False,
        "has_cuda_kernel": False,
        "has_cpu_kernel": False,
    }

    # Parse op name: aten::name.overload -> name, overload
    name_part = op_name
    if name_part.startswith("aten::"):
        name_part = name_part[6:]

    base_name = name_part.split(".")[0]
    overload = name_part.split(".")[1] if "." in name_part else "default"

    try:
        # Try to get the operator from torch.ops.aten
        op_obj = getattr(torch.ops.aten, base_name, None)
        if op_obj is None:
            return info

        # Get the specific overload
        try:
            op_overload = getattr(op_obj, overload, None)
            if op_overload is None:
                op_overload = op_obj.default if hasattr(op_obj, 'default') else None
        except Exception:
            op_overload = None

        if op_overload is None:
            return info

        # Check dispatch keys
        try:
            dispatch_keys = op_overload._dispatch_keys()
            if dispatch_keys:
                keys_str = str(dispatch_keys)
                info["is_composite_implicit"] = "CompositeImplicitAutograd" in keys_str
                info["is_composite_explicit"] = "CompositeExplicitAutograd" in keys_str
                info["has_cuda_kernel"] = "CUDA" in keys_str
                info["has_cpu_kernel"] = "CPU" in keys_str
        except Exception:
            pass

        # Alternative: try _dispatch_has_kernel_for_dispatch_key
        if not any(info.values()):
            for key, attr in [
                ("is_composite_implicit", "CompositeImplicitAutograd"),
                ("is_composite_explicit", "CompositeExplicitAutograd"),
                ("has_cuda_kernel", "CUDA"),
                ("has_cpu_kernel", "CPU"),
            ]:
                try:
                    result = torch._C._dispatch_has_kernel_for_dispatch_key(
                        op_overload.name(), attr
                    )
                    info[key] = result
                except Exception:
                    pass

    except Exception as e:
        pass

    return info


# ============================================================
# 4. Triton lowering detection
# ============================================================

def get_triton_lowered_ops():
    """Get the set of ops that have Triton lowerings in torch._inductor."""
    triton_ops = set()
    fallback_ops = set()

    try:
        # Import inductor lowering module
        import torch._inductor.lowering as lowering_mod

        # Get lowerings dict
        if hasattr(lowering_mod, 'lowerings'):
            for key in lowering_mod.lowerings:
                try:
                    name = str(key)
                    if "aten" in name:
                        triton_ops.add(key)
                except Exception:
                    pass

        # Get fallback ops (these are NOT triton-lowered, they fall back to eager)
        if hasattr(lowering_mod, 'fallbacks'):
            for key in lowering_mod.fallbacks:
                try:
                    fallback_ops.add(key)
                except Exception:
                    pass

    except ImportError:
        print("Warning: torch._inductor not available, Triton detection skipped.")
    except Exception as e:
        print(f"Warning: Error loading inductor lowerings: {e}")

    return triton_ops, fallback_ops


def has_triton_lowering(op_name: str, triton_ops: set, fallback_ops: set) -> bool:
    """Check if an operator has a Triton lowering."""
    name_part = op_name
    if name_part.startswith("aten::"):
        name_part = name_part[6:]

    base_name = name_part.split(".")[0]
    overload = name_part.split(".")[1] if "." in name_part else "default"

    # Try to match against the triton_ops set
    try:
        op_obj = getattr(torch.ops.aten, base_name, None)
        if op_obj is None:
            return False
        op_overload = getattr(op_obj, overload, None)
        if op_overload is None and overload == "default":
            op_overload = getattr(op_obj, "default", None)
        if op_overload is None:
            return False

        # Check if this exact op is in triton lowerings but not in fallbacks
        if op_overload in triton_ops and op_overload not in fallback_ops:
            return True

        # Also check by name matching
        for t_op in triton_ops:
            try:
                t_name = t_op.name() if hasattr(t_op, 'name') else str(t_op)
                if base_name in str(t_name):
                    if t_op not in fallback_ops:
                        return True
            except Exception:
                pass

    except Exception:
        pass

    return False


# ============================================================
# 5. Public API name mapping
# ============================================================

def build_public_api_mapping():
    """Build mapping from aten op base name to public API names."""
    mapping = defaultdict(set)

    modules_to_scan = [
        ("torch", torch),
        ("torch.nn.functional", None),
        ("torch.linalg", None),
        ("torch.fft", None),
        ("torch.special", None),
    ]

    for mod_name, mod_obj in modules_to_scan:
        if mod_obj is None:
            try:
                parts = mod_name.split(".")
                mod_obj = torch
                for p in parts[1:]:
                    mod_obj = getattr(mod_obj, p)
            except AttributeError:
                continue

        for attr_name in dir(mod_obj):
            if attr_name.startswith("__"):
                continue
            try:
                obj = getattr(mod_obj, attr_name)
            except Exception:
                continue

            if not callable(obj):
                continue

            public_name = f"{mod_name}.{attr_name}"

            # The attr_name often matches the aten op base name
            mapping[attr_name].add(public_name)

            # Also try to find decomposition info
            if hasattr(obj, '__name__'):
                mapping[obj.__name__].add(public_name)

    return mapping


# ============================================================
# 6. Tag detection
# ============================================================

def get_op_tags(op_name: str) -> str:
    """Try to get tags for an operator."""
    name_part = op_name
    if name_part.startswith("aten::"):
        name_part = name_part[6:]

    base_name = name_part.split(".")[0]
    overload = name_part.split(".")[1] if "." in name_part else "default"

    tags = []
    try:
        op_obj = getattr(torch.ops.aten, base_name, None)
        if op_obj is None:
            return ""
        op_overload = getattr(op_obj, overload, None)
        if op_overload is None and overload == "default":
            op_overload = getattr(op_obj, "default", None)
        if op_overload is None:
            return ""

        if hasattr(op_overload, 'tags'):
            raw_tags = op_overload.tags
            if raw_tags:
                for t in raw_tags:
                    tags.append(str(t).replace("torch.Tag.", ""))
    except Exception:
        pass

    return ", ".join(tags)


# ============================================================
# 7. Classification logic
# ============================================================

CATEGORY_NAMES = {
    1: "CUDA直接提供算子",
    2: "CUDA C++自定义算子",
    3: "算子拼接",
    4: "Triton开发",
}


def classify_op(op_name: str, dispatch_info: dict, has_triton: bool) -> int:
    """
    Classify an operator according to priority:
    1. Composite (算子拼接) -> 3
    2. Triton lowering -> 4
    3. CUDA library op -> 1
    4. Remaining per-backend kernel -> 2
    """
    base_name = normalize_op_name(op_name)

    # Priority 1: Composite ops
    if dispatch_info["is_composite_implicit"] or dispatch_info["is_composite_explicit"]:
        return 3

    # Priority 2: Triton
    if has_triton and not is_cuda_library_op(base_name):
        return 4

    # Priority 3: CUDA library
    if is_cuda_library_op(base_name):
        return 1

    # Priority 4: Custom CUDA C++ kernel
    if dispatch_info["has_cuda_kernel"] or dispatch_info["has_cpu_kernel"]:
        return 2

    # Default: if we can't determine, mark as CUDA C++ custom
    return 2


# ============================================================
# Main
# ============================================================

def main():
    print(f"PyTorch version: {torch.__version__}")
    print(f"CUDA available: {torch.cuda.is_available()}")
    print()

    # Step 1: Get all aten ops
    print("Enumerating ATen operators...")
    aten_ops = get_all_aten_ops()
    print(f"  Found {len(aten_ops)} ATen operator schemas")

    # Step 2: Get Triton lowering info
    print("Loading Triton lowering information...")
    triton_ops, fallback_ops = get_triton_lowered_ops()
    print(f"  Triton lowerings: {len(triton_ops)}, Fallbacks: {len(fallback_ops)}")

    # Step 3: Build public API mapping
    print("Building public API name mapping...")
    api_mapping = build_public_api_mapping()
    print(f"  Mapped {len(api_mapping)} base names to public APIs")

    # Step 4: Classify each operator
    print("Classifying operators...")
    results = []
    category_counts = defaultdict(int)

    for op_info in aten_ops:
        full_name = op_info["full_name"]
        base_name = normalize_op_name(full_name)

        # Get dispatch info
        dispatch_info = get_op_dispatch_info(full_name)

        # Check Triton
        triton = has_triton_lowering(full_name, triton_ops, fallback_ops)

        # Classify
        category = classify_op(full_name, dispatch_info, triton)
        category_counts[category] += 1

        # Public API names
        public_names = api_mapping.get(base_name, set())
        public_name_str = ", ".join(sorted(public_names)) if public_names else ""

        # Tags
        tags = get_op_tags(full_name)

        results.append({
            "算子名称": full_name,
            "公共API名称": public_name_str,
            "分类编号": category,
            "分类名称": CATEGORY_NAMES[category],
            "是否有Triton lowering": "是" if triton else "否",
            "Tags": tags,
        })

    # Print summary
    print("\nClassification Summary:")
    for cat in sorted(CATEGORY_NAMES.keys()):
        print(f"  {cat}-{CATEGORY_NAMES[cat]}: {category_counts[cat]}")
    print(f"  Total: {sum(category_counts.values())}")

    # Step 5: Write to Excel
    output_path = r"D:\work\PyTorch\tools\torch_API_classification\pytorch_api_classification.xlsx"
    print(f"\nWriting results to {output_path}...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PyTorch API Classification"

    # Header
    headers = ["算子名称", "公共API名称", "分类编号", "分类名称", "是否有Triton lowering", "Tags"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data rows
    for row_idx, result in enumerate(results, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=result[header])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(output_path)
    print(f"Done! Saved {len(results)} operators to Excel.")

    # Print a few examples for verification
    print("\nSample results for verification:")
    samples = ["aten::mm", "aten::conv2d", "aten::relu", "aten::add.Tensor",
               "aten::fft_fft", "aten::linalg_svd"]
    for s in samples:
        for r in results:
            if r["算子名称"] == s:
                print(f"  {r['算子名称']}: {r['分类编号']}-{r['分类名称']}, "
                      f"Triton={r['是否有Triton lowering']}, Tags={r['Tags']}")
                break


if __name__ == "__main__":
    main()
