from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import List

from .finder import SymbolResolution, resolve_symbols

DEFAULT_SYMBOLS = [
    "torch._C._cuda_attach_out_of_memory_observer",
    "torch._C._cuda_beginAllocateCurrentThreadToPool",
    "torch._C._cuda_clearCublasWorkspaces",
    "torch._C._cuda_endAllocateToPool",
    "torch._C._cuda_getDeviceCount",
    "torch._C._cuda_setRNGState",
    "torch._C._dynamo.eval_frame._set_lru_cache",
    "torch._dynamo.functional_export.dynamo_graph_capture_for_export",
    "torch._dynamo.is_exporting",
    "torch._grouped_mm",
    "torch._inductor.utils.is_cudagraph_unsafe_op",
    "torch.accelerator.default_stream",
    "torch.accelerator.Event",
    "torch.accelerator.manual_seed",
    "torch.accelerator.manual_seed_all",
    "torch.accelerator.stream",
    "torch.cuda.amp.grad_scaler._MultiDeviceReplicator",
    "torch.cuda.default_generators.seed",
    "torch.distributed._composable.fsdp.CPUOffloadPolicy",
    "torch.distributed._composable.fsdp.fully_shard",
    "torch.distributed._composable.fsdp.MixedPrecisionPolicy",
    "torch.distributed._composable.fsdp.OffloadPolicy",
    "torch.distributed._composable.replicate.replicate",
    "torch.distributed._mesh_layout._MeshLayout",
    "torch.distributed._symmetric_memory.set_backend",
    "torch.distributed._tensor._utils.compute_local_shape_and_global_offset",
    "torch.distributed._tools.fsdp2_mem_tracker.FSDPMemTracker",
    "torch.distributed.algorithms.ddp_comm_hooks.post_localSGD_hook.PostLocalSGDState",
    "torch.distributed.checkpoint._consolidate_hf_safetensors.consolidate_safetensors_files_on_every_rank",
    "torch.distributed.checkpoint.format_utils._EmptyStateDictLoadPlanner",
    "torch.distributed.checkpoint.format_utils._load_state_dict",
    "torch.distributed.checkpoint.format_utils.dcp_to_torch_save",
    "torch.distributed.checkpoint.HuggingFaceStorageReader",
    "torch.distributed.checkpoint.HuggingFaceStorageWriter",
    "torch.distributed.checkpoint.quantized_hf_storage.QuantizedHuggingFaceStorageReader",
    "torch.distributed.checkpoint.staging.DefaultStager",
    "torch.distributed.checkpoint.staging.StagingOptions",
    "torch.distributed.checkpoint.state_dict.get_model_state_dict",
    "torch.distributed.checkpoint.state_dict.get_optimizer_state_dict",
    "torch.distributed.checkpoint.state_dict.get_state_dict",
    "torch.distributed.checkpoint.state_dict.set_model_state_dict",
    "torch.distributed.checkpoint.state_dict.set_optimizer_state_dict",
    "torch.distributed.checkpoint.state_dict.set_state_dict",
    "torch.distributed.checkpoint.state_dict.StateDictOptions",
    "torch.distributed.device_mesh.DeviceMesh._concatenate",
    "torch.distributed.distributed_c10d._all_gather_base_coalesced",
    "torch.distributed.distributed_c10d._shutdown_backend",
    "torch.distributed.distributed_c10d.ProcessGroupXCCL",
    "torch.distributed.fetch",
    "torch.distributed.pipelining.PipelineStage",
    "torch.distributed.pipelining.ScheduleGPipe",
    "torch.distributed.pipelining.schedules.get_schedule_class",
    "torch.distributed.ProcessGroup.Options",
    "torch.distributed.ProcessGroupNCCL.Options",
    "torch.distributed.rendezvous.rendezvous",
    "torch.distributed.ring_exchange",
    "torch.distributed.tensor.parallel.PrepareModuleInputOutput",
    "torch.get_current_dtype",
    "torch.multiprocessing.get_start_methods",
    "torch.multiprocessing.reductions._rebuild_cuda_tensor_original",
    "torch.multiprocessing.reductions._reduce_tensor_original",
    "torch.nested._internal.ops.extract_kwargs",
    "torch.nn.attention.flex_attention.and_masks",
    "torch.nn.attention.flex_attention.BlockMask.from_kv_blocks",
    "torch.nn.attention.flex_attention.create_block_mask",
    "torch.nn.attention.flex_attention.flex_attention",
    "torch.nn.nn.LayerNorm",
    "torch.nn.utils.spectral_norm.SpectralNorm.apply",
    "torch.optim.adam.Adam",
    "torch.optim.adamw.AdamW",
    "torch.optim.optimizer._default_to_fused_or_foreach",
    "torch.optim.optimizer._disable_dynamo_if_unsupported",
    "torch.optim.optimizer._get_capturable_supported_devices",
    "torch.optim.optimizer._get_value",
    "torch.optim.optimizer._stack_if_compiling",
    "torch.optim.optimizer._view_as_real",
    "torch.optim.optimizer.Optimizer",
    "torch.optim.optimizer.Optimizer._group_tensors_by_device_and_dtype",
    "torch.random.randint",
    "torch.version.cuda.split",
    "torch.zeroes",
]

def _read_symbols_from_file(path: Path) -> List[str]:
    txt = path.read_text(encoding="utf-8")
    syms: List[str] = []
    for line in txt.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        # allow comma separated
        for part in line.split(","):
            p = part.strip()
            if p:
                syms.append(p)
    return syms

def _build_row(r: SymbolResolution):
    """Return (original_api, import_format, extra_info, status) for one result."""
    if r.status == "ok":
        if r.attrpath:
            import_fmt = f"from {r.module} import {r.attrpath[0]}"
            extra = f".{'.'.join(r.attrpath[1:])}" if len(r.attrpath) > 1 else ""
        else:
            import_fmt = f"import {r.module}"
            extra = ""
        if r.hint_modules:
            status = "HINT"
        else:
            status = "OK"
    else:
        import_fmt = ""
        extra = ""
        if r.hint_modules:
            extra = ", ".join(r.hint_modules[:8])
            status = "HINT"
        else:
            status = "MISS"
    return r.symbol, import_fmt, extra, status


def _write_excel(results: List[SymbolResolution], path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "torch_symbols"

    headers = ["原始接口", "正确import格式", "备注", "状态"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_fills = {
        "OK": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "MISS": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "HINT": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }

    for idx, r in enumerate(results, 2):
        symbol, import_fmt, extra, status = _build_row(r)
        ws.cell(row=idx, column=1, value=symbol).border = thin_border
        ws.cell(row=idx, column=2, value=import_fmt).border = thin_border
        ws.cell(row=idx, column=3, value=extra).border = thin_border
        sc = ws.cell(row=idx, column=4, value=status)
        sc.border = thin_border
        sc.alignment = Alignment(horizontal="center")
        if status in status_fills:
            sc.fill = status_fills[status]

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 10

    wb.save(path)


def main(argv: List[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        prog="torch-symbol-finder",
        description="Resolve dotted torch symbols to the real import paths available in your current environment.",
    )
    p.add_argument(
        "--symbols-file",
        type=str,
        default=None,
        help="Path to a text file containing dotted symbols (one per line; # comments allowed).",
    )
    p.add_argument(
        "--json",
        action="store_true",
        help="Output JSON (machine readable).",
    )
    p.add_argument(
        "--no-hints",
        action="store_true",
        help="Disable __all__ scanning for hint modules (faster).",
    )
    p.add_argument(
        "--excel",
        type=str,
        default=None,
        help="Output results to an Excel file (.xlsx).",
    )
    args = p.parse_args(argv)

    if args.symbols_file:
        symbols = _read_symbols_from_file(Path(args.symbols_file))
    else:
        symbols = DEFAULT_SYMBOLS

    # Basic environment info
    env = {}
    try:
        import torch
        env = {"torch_version": getattr(torch, "__version__", None), "torch_cuda_version": getattr(torch.version, "cuda", None)}
    except Exception as e:
        env = {"torch_import_error": str(e)}

    results = resolve_symbols(symbols, with_hints=not args.no_hints)

    if args.excel:
        _write_excel(results, args.excel)
        sys.stdout.write(f"Results written to {args.excel}\n")
        return 0

    if args.json:
        payload = {
            "env": env,
            "results": [r.as_dict() for r in results],
        }
        sys.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")
        return 0

    # Pretty text output
    sys.stdout.write(f"torch.__version__ = {env.get('torch_version')}\n")
    sys.stdout.write(f"torch.version.cuda = {env.get('torch_cuda_version')}\n\n")
    for r in results:
        if r.status == "ok":
            attr = ".".join(r.attrpath or [])
            if r.attrpath:
                sys.stdout.write(f"[OK]   {r.symbol}\n       => from {r.module} import {r.attrpath[0]}")
                if len(r.attrpath) > 1:
                    sys.stdout.write(f"  (then .{'.'.join(r.attrpath[1:])})")
                sys.stdout.write("\n")
            else:
                sys.stdout.write(f"[OK]   {r.symbol}\n       => import {r.module}\n")
        else:
            sys.stdout.write(f"[MISS] {r.symbol}\n")
            if r.hint_modules:
                sys.stdout.write(f"       hints: {', '.join(r.hint_modules[:8])}\n")
            if r.error:
                sys.stdout.write(f"       error: {r.error}\n")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
