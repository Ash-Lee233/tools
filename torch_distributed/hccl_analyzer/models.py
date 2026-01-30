"""Data structures for HCCL dependency analysis."""

from __future__ import annotations

import enum
from dataclasses import dataclass, field
from typing import Any


class Confidence(enum.Enum):
    """Confidence tier for an HCCL dependency."""

    HIGH = "high"      # Tier 1: direct HCCL reference in same file
    MEDIUM = "medium"  # Tier 2: via ProcessGroup abstraction
    LOW = "low"        # Tier 3: multi-hop call chain (BFS depth <= 3)


@dataclass
class HCCLReference:
    """A single occurrence of an HCCL-related pattern in source code."""

    file: str
    line: int
    pattern: str   # the matched pattern string
    context: str   # the source line containing the match


@dataclass
class DistributedAPI:
    """A public API discovered under ``torch.distributed``."""

    qualified_name: str   # e.g. "torch.distributed.all_reduce"
    file: str
    line: int
    kind: str = "function"  # function | class | method


@dataclass
class DependencyEdge:
    """One hop in a dependency chain from an API to an HCCL reference."""

    caller: str   # qualified name or file:line
    callee: str
    file: str
    line: int


@dataclass
class AnalysisResult:
    """Result for a single API that depends on HCCL."""

    api: DistributedAPI
    confidence: Confidence
    references: list[HCCLReference] = field(default_factory=list)
    chain: list[DependencyEdge] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "api": self.api.qualified_name,
            "file": self.api.file,
            "line": self.api.line,
            "kind": self.api.kind,
            "confidence": self.confidence.value,
            "references": [
                {"file": r.file, "line": r.line, "pattern": r.pattern, "context": r.context}
                for r in self.references
            ],
            "chain": [
                {"caller": e.caller, "callee": e.callee, "file": e.file, "line": e.line}
                for e in self.chain
            ],
        }


@dataclass
class AnalysisReport:
    """Complete analysis report."""

    repo_path: str
    total_apis: int = 0
    results: list[AnalysisResult] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        grouped: dict[str, list[dict[str, Any]]] = {"high": [], "medium": [], "low": []}
        for r in self.results:
            grouped[r.confidence.value].append(r.to_dict())
        return {
            "repo_path": self.repo_path,
            "total_apis_scanned": self.total_apis,
            "hccl_dependent_apis": len(self.results),
            "by_confidence": grouped,
        }

    def format_text(self) -> str:
        lines: list[str] = []
        lines.append("=" * 70)
        lines.append("HCCL Dependency Analysis Report")
        lines.append("=" * 70)
        lines.append(f"Repository : {self.repo_path}")
        lines.append(f"Total APIs scanned : {self.total_apis}")
        lines.append(f"HCCL-dependent APIs: {len(self.results)}")
        lines.append("")

        for tier, label in [
            (Confidence.HIGH, "Tier 1 - HIGH confidence (direct HCCL reference)"),
            (Confidence.MEDIUM, "Tier 2 - MEDIUM confidence (via ProcessGroup abstraction)"),
            (Confidence.LOW, "Tier 3 - LOW confidence (multi-hop call chain)"),
        ]:
            items = [r for r in self.results if r.confidence == tier]
            lines.append("-" * 70)
            lines.append(f"{label}  [{len(items)} APIs]")
            lines.append("-" * 70)
            if not items:
                lines.append("  (none)")
                lines.append("")
                continue
            for r in items:
                lines.append(f"  {r.api.qualified_name}")
                lines.append(f"    defined at {r.api.file}:{r.api.line}  ({r.api.kind})")
                if r.references:
                    lines.append("    HCCL references:")
                    for ref in r.references:
                        lines.append(f"      - {ref.file}:{ref.line}  pattern={ref.pattern!r}")
                        lines.append(f"        {ref.context.strip()}")
                if r.chain:
                    lines.append("    call chain:")
                    for edge in r.chain:
                        lines.append(f"      {edge.caller} -> {edge.callee}  ({edge.file}:{edge.line})")
                lines.append("")

        return "\n".join(lines)

    def export_excel(self, path: str) -> None:
        """Write the report to an Excel file."""
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "HCCL Dependencies"

        # Column widths
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 80

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        wrap_align = Alignment(wrap_text=True, vertical="top")

        # Write headers
        headers = ["Tier", "API", "File", "References"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_align

        # Tier label mapping
        tier_labels = {
            Confidence.HIGH: "Tier 1 (HIGH)",
            Confidence.MEDIUM: "Tier 2 (MEDIUM)",
            Confidence.LOW: "Tier 3 (LOW)",
        }

        # Tier row fills for visual grouping
        tier_fills = {
            Confidence.HIGH: PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
            Confidence.MEDIUM: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            Confidence.LOW: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        }

        def _rel(filepath: str) -> str:
            """Convert absolute path to relative path from repo_path."""
            try:
                return os.path.relpath(filepath, self.repo_path)
            except ValueError:
                return filepath

        row = 2
        for result in self.results:
            tier_label = tier_labels[result.confidence]
            api_name = result.api.qualified_name
            file_loc = f"{_rel(result.api.file)}:{result.api.line}"
            refs_text = self._format_refs_for_excel(result, _rel)

            fill = tier_fills[result.confidence]
            for col, value in enumerate([tier_label, api_name, file_loc, refs_text], start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = wrap_align
                cell.fill = fill

            row += 1

        wb.save(path)

    @staticmethod
    def _format_refs_for_excel(
        result: "AnalysisResult",
        rel_path: "Any",
    ) -> str:
        """Build a concise, human-readable references string for Excel."""
        # Pattern -> human-readable label
        _PATTERN_LABELS: dict[str, str] = {
            r"\bProcessGroupHCCL\b": "ProcessGroupHCCL",
            r"""['"]hccl['"]""": '"hccl"',
            r"\bHCCL\b": "HCCL",
            r"\bhcom\b": "hcom",
            r"\bProcessGroupNPU\b": "ProcessGroupNPU",
            r"\bhccl_available\b": "hccl_available",
        }

        lines: list[str] = []

        if result.confidence == Confidence.LOW:
            # Tier 3: show call chain only
            if result.chain:
                chain_str = " -> ".join(e.callee for e in result.chain)
                lines.append(f"{result.chain[0].caller} -> {chain_str}")
            return "\n".join(lines)

        if result.confidence == Confidence.MEDIUM:
            # Tier 2: show PG dispatch calls
            seen: set[str] = set()
            for ref in result.references:
                ctx = ref.context.strip()
                if ctx not in seen:
                    seen.add(ctx)
                    lines.append(f"L{ref.line}: {ctx}")
            return "\n".join(lines)

        # Tier 1: filter to key references, deduplicate, cap at 5
        # Priority: "hccl" string / ProcessGroupHCCL / HCCL keyword > generic imports
        key_refs: list[HCCLReference] = []
        generic_refs: list[HCCLReference] = []
        for ref in result.references:
            if ref.pattern in _PATTERN_LABELS:
                key_refs.append(ref)
            else:
                generic_refs.append(ref)

        # If no key refs, fall back to generic ones
        chosen = key_refs if key_refs else generic_refs

        seen_ctx: set[str] = set()
        for ref in chosen:
            ctx = ref.context.strip()
            if ctx in seen_ctx:
                continue
            seen_ctx.add(ctx)
            label = _PATTERN_LABELS.get(ref.pattern, ref.pattern)
            lines.append(f"L{ref.line} [{label}]: {ctx}")
            if len(lines) >= 5:
                break

        remaining = len(chosen) - len(lines)
        if remaining > 0:
            lines.append(f"... +{remaining} more")

        return "\n".join(lines)
